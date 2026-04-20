"""Generate Store Caller Test Cases Excel file."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Store Caller Tests"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
scenario_font = Font(bold=True, size=11)
scenario_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Headers
headers = [
    "Test #",
    "Scenario",
    "Test Case",
    "Priority",
    "Fix #",
    "Caller Says",
    "Expected Behavior",
    "Pass/Fail",
    "Actual Behavior",
    "Notes",
]
col_widths = [8, 30, 40, 10, 8, 35, 55, 10, 40, 30]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Test data: (scenario, test_case, priority, fix_num, caller_says, expected)
tests = [
    # Scenario 1: Retailer with PO Number (Happy Path)
    ("1. Retailer + PO (Happy Path)", "Retailer identifies and gives PO number", "P0", "2",
     "\"I'm from Lowe's\" then \"PO number is 523824\"",
     "Asks for PO/project#. Calls ask_store_bot with ALL 3 fields: question, lookup_type=po_number, lookup_value=523824"),
    ("", "Returns project status by type, not number", "P0", "5",
     "(after tool returns project data)",
     "Says 'your flooring installation is ready to schedule' — NOT 'project 74356 is ready'"),
    ("", "Follow-up question uses session", "P1", "3",
     "\"Is a technician assigned?\"",
     "Calls ask_store_bot with just question. System remembers project context. Returns real data."),
    ("", "Ends call politely", "P2", "9",
     "\"That's all, thanks\"",
     "Ends call politely. No mention of 'transfer'. No abrupt hangup."),

    # Scenario 2: Retailer with Project Number
    ("2. Retailer + Project Number", "STT digits with spaces stripped", "P0", "2",
     "\"7 4 3 5 6\"",
     "Strips spaces, passes '74356' to ask_store_bot with lookup_type=project_number"),
    ("", "Bad project number — no fabrication", "P0", "1",
     "\"789\"",
     "Tool returns 'not found'. Says 'I couldn't find a project with that number.' Does NOT invent status/dates/technician."),
    ("", "Second attempt succeeds", "P0", "2",
     "\"74356\"",
     "Authenticates successfully, returns real project data from tool."),

    # Scenario 3: Wrong Identifier Rejected
    ("3. Wrong Identifier Rejected", "Rejects customer name", "P0", "2",
     "\"It's for John Smith\"",
     "\"I can only look up projects by project number or PO number. Do you have either of those?\""),
    ("", "Rejects address", "P0", "2",
     "\"123 Main Street\"",
     "Same rejection — only PO or project number accepted."),
    ("", "Rejects phone number", "P0", "2",
     "\"555-123-4567\"",
     "Same rejection — only PO or project number accepted."),
    ("", "Rejects project description", "P0", "2",
     "\"The flooring job at the Smith house\"",
     "Same rejection — only PO or project number accepted."),
    ("", "Eventually accepts valid PO", "P0", "2",
     "\"OK, PO is 523824\"",
     "Proceeds with auth normally. All 3 fields passed to ask_store_bot."),

    # Scenario 4: Scheduling Blocked
    ("4. Scheduling Blocked for Retailers", "Asks to schedule after status", "P1", "—",
     "\"Can I schedule the installation?\"",
     "\"Scheduling is not available for retailer calls. Please have the customer call us directly.\""),
    ("", "Asks how to proceed", "P1", "9",
     "\"How do I get this scheduled?\"",
     "Same — direct customer to call. Does NOT say 'transferring' or 'let me connect you'."),
    ("", "Pushes back", "P1", "—",
     "\"Can't you just do it?\"",
     "Firm but polite — retailers cannot schedule. Does not cave or fabricate a booking."),

    # Scenario 5: Transfer Available
    ("5. Transfer Available (support# exists)", "Customer identified — transfers", "P0", "9",
     "\"I'm a customer\"",
     "\"I don't recognize your phone number. Let me transfer you.\" Then ACTUALLY invokes transferCall tool."),
    ("", "Non-project call — transfers", "P1", "9",
     "\"I want a job application\"",
     "\"Let me connect you with someone who can help.\" Then invokes transferCall tool."),
    ("", "Retailer needs human help", "P1", "9",
     "\"I need to speak to someone\"",
     "Offers transfer via transferCall tool. Call actually transfers."),

    # Scenario 6: NO Transfer Available
    ("6. No Transfer (support# empty)", "Customer identified — no transfer promise", "P0", "9",
     "\"I'm a customer\"",
     "\"I don't have your account on file. Our team at Projects Force will reach out shortly. Anything else?\" — NEVER says 'transferring' or 'let me connect you'."),
    ("", "Non-project call — no transfer promise", "P0", "9",
     "\"I need help with billing\"",
     "\"Our team at Projects Force will reach out shortly.\" — NO 'transferring', NO 'let me connect you'."),
    ("", "Caller says bye — no premature end", "P2", "7",
     "\"OK bye\"",
     "Ends call politely. Does NOT hang up mid-sentence. (bare 'bye' removed from endCallPhrases)"),
    ("", "No stuck transfer loop", "P0", "9",
     "(any dead end)",
     "NEVER repeats 'I'm transferring you now'. Says 'reach out' once, asks if anything else, ends call."),

    # Scenario 7: Confused/Vague Caller
    ("7. Confused / Vague Caller", "\"What do you do?\"", "P1", "—",
     "\"What do you do?\"",
     "Gives orientation: 'I help retailers check project status and help customers schedule appointments. Are you a customer or retailer?'"),
    ("", "\"I don't know anything\"", "P1", "9",
     "\"I don't know anything\"",
     "Same orientation — does NOT immediately transfer or end call. Gives 2 attempts."),
    ("", "Still confused after 2 tries", "P2", "9",
     "\"I still don't understand\" then \"Help me\"",
     "After 2 attempts: 'Our team will reach out' (no transfer) or transfers (if available). Does NOT loop."),
    ("", "Irrelevant request", "P2", "9",
     "\"I want water\"",
     "Non-project path. 'Our team will reach out' or transfer. Does NOT just hang up."),

    # Scenario 8: Filler Behavior
    ("8. Filler Behavior", "New question — single filler", "P1", "4",
     "\"What's the status of my project?\"",
     "Says 'One moment.' ONCE, then calls tool. No other filler."),
    ("", "Reply to question — no filler", "P1", "4",
     "Bot asks 'customer or retailer?' — caller says 'Retailer'",
     "NO filler. Goes straight to 'Could you provide the project number or PO number?'"),
    ("", "Forbidden fillers never used", "P1", "4",
     "(any tool call)",
     "NEVER says: 'Hold on', 'Wait', 'Hang on', 'Just a sec', 'Let me pull that up', 'One second', 'Let me take a look'."),

    # Scenario 9: Caller Needs Time
    ("9. Caller Needs Time", "\"Just a second\"", "P2", "7",
     "\"Just a second, let me find the number\"",
     "\"Take your time, I'll be right here.\" Waits patiently. Does NOT end call."),
    ("", "Long pause under 60s", "P2", "7",
     "(45 seconds silence after 'just a second')",
     "Does NOT time out. 60s silence timeout gives them time."),
    ("", "Caller says 'hold on'", "P2", "7",
     "\"Hold on\"",
     "Waits patiently. Does NOT treat as end-call phrase. Does NOT hang up."),

    # Scenario 10: Hallucination Prevention
    ("10. Hallucination Prevention", "Tool returns vague response", "P0", "1",
     "(tool returns 'I need a project number')",
     "Tells caller exactly that. Does NOT invent status, dates, technician names, or any details."),
    ("", "Tool returns real data", "P0", "1",
     "(tool returns project details)",
     "Reads ONLY what tool returned. No extra details, no embellishments."),
    ("", "Asks about field not in response", "P0", "1",
     "\"What's the customer's phone number?\"",
     "\"I can't share customer contact details\" or \"I don't have that information\". Does NOT fabricate."),

    # Scenario 11: TTS / Speech Quality
    ("11. TTS / Speech Quality", "Company name pronunciation", "P2", "6",
     "(any mention of company name)",
     "Says 'Projects Force' (two words). NOT 'ProjectsForce' (one word) or 'Project Source'."),
    ("", "Project numbers NOT read aloud", "P2", "5",
     "(after tool returns project data)",
     "Says 'your flooring installation' or 'your project'. NEVER reads '7 4 3 5 6' aloud."),
    ("", "End call message clear", "P2", "6",
     "(call ends)",
     "\"Thank you for calling Projects Force. Have a great day!\" — clear pronunciation."),
]

# Write data
row = 2
test_num = 1
for scenario, test_case, priority, fix_num, caller_says, expected in tests:
    ws.cell(row=row, column=1, value=f"TC-{test_num:02d}").alignment = wrap
    ws.cell(row=row, column=1).border = thin_border

    cell_scenario = ws.cell(row=row, column=2, value=scenario)
    cell_scenario.alignment = wrap
    cell_scenario.border = thin_border
    if scenario:
        cell_scenario.font = scenario_font
        cell_scenario.fill = scenario_fill

    ws.cell(row=row, column=3, value=test_case).alignment = wrap
    ws.cell(row=row, column=3).border = thin_border

    cell_pri = ws.cell(row=row, column=4, value=priority)
    cell_pri.alignment = wrap
    cell_pri.border = thin_border
    if priority == "P0":
        cell_pri.font = Font(bold=True, color="CC0000")
    elif priority == "P1":
        cell_pri.font = Font(bold=True, color="E67E00")

    ws.cell(row=row, column=5, value=fix_num).alignment = wrap
    ws.cell(row=row, column=5).border = thin_border

    ws.cell(row=row, column=6, value=caller_says).alignment = wrap
    ws.cell(row=row, column=6).border = thin_border

    ws.cell(row=row, column=7, value=expected).alignment = wrap
    ws.cell(row=row, column=7).border = thin_border

    ws.cell(row=row, column=8, value="").alignment = wrap
    ws.cell(row=row, column=8).border = thin_border

    ws.cell(row=row, column=9, value="").alignment = wrap
    ws.cell(row=row, column=9).border = thin_border

    ws.cell(row=row, column=10, value="").alignment = wrap
    ws.cell(row=row, column=10).border = thin_border

    row += 1
    test_num += 1

# Add data validation for Pass/Fail column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Skipped"', allow_blank=True)
dv.error = "Please select Pass, Fail, Blocked, or Skipped"
dv.errorTitle = "Invalid entry"
ws.add_data_validation(dv)
for r in range(2, row):
    dv.add(ws.cell(row=r, column=8))

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:J{row - 1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 15

summary_data = [
    ("Total Test Cases", test_num - 1, ""),
    ("", "", ""),
    ("By Priority", "Count", ""),
    ("P0 (Critical)", sum(1 for t in tests if t[2] == "P0"), ""),
    ("P1 (High)", sum(1 for t in tests if t[2] == "P1"), ""),
    ("P2 (Medium)", sum(1 for t in tests if t[2] == "P2"), ""),
    ("", "", ""),
    ("By Scenario", "Count", ""),
    ("1. Retailer + PO (Happy Path)", 4, ""),
    ("2. Retailer + Project Number", 3, ""),
    ("3. Wrong Identifier Rejected", 5, ""),
    ("4. Scheduling Blocked", 3, ""),
    ("5. Transfer Available", 3, ""),
    ("6. No Transfer Available", 4, ""),
    ("7. Confused / Vague Caller", 4, ""),
    ("8. Filler Behavior", 3, ""),
    ("9. Caller Needs Time", 3, ""),
    ("10. Hallucination Prevention", 3, ""),
    ("11. TTS / Speech Quality", 3, ""),
    ("", "", ""),
    ("Test Date", "", ""),
    ("Tester", "", ""),
    ("Environment", "QA / Dev / Prod", ""),
    ("Build / Commit", "", ""),
]

for r, (a, b, c) in enumerate(summary_data, 1):
    ws2.cell(row=r, column=1, value=a).font = Font(bold=True) if r in (1, 3, 8) else Font()
    ws2.cell(row=r, column=2, value=b)
    ws2.cell(row=r, column=3, value=c)

out = "/Users/jjayaraj/workspaces/studios/projectsforce/schedulingAIBot/docs/Store_Caller_Test_Cases.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total test cases: {test_num - 1}")
