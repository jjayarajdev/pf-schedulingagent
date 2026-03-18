"""Integration test configuration — credential input, HTTP capture, Excel report.

Credential resolution order:
  1. pytest CLI options:  --pf-email, --pf-password, --pf-url, --pf-identifier
  2. Environment variables: PF_TEST_EMAIL, PF_TEST_PASSWORD, PF_INTEGRATION_URL, PF_TEST_IDENTIFIER
  3. Config file:          tests/integration/.pf-creds.json
  4. Interactive prompt:   (if running in a terminal)

HTTP Capture:
  All httpx requests/responses are captured and written to an Excel report at
  the end of the test session: ``E2E_Scheduling_Test_Report.xlsx``
"""

import json
import logging
import os
import sys
from datetime import datetime, timezone
from unittest.mock import AsyncMock

import httpx
import pytest

# Ensure src/ is on the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src"))

logger = logging.getLogger(__name__)

_PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
_CREDS_FILE = os.path.join(os.path.dirname(__file__), ".pf-creds.json")
_REPORT_FILE = os.path.join(_PROJECT_ROOT, "E2E_Scheduling_Test_Report.xlsx")


# ---------------------------------------------------------------------------
#  pytest CLI options
# ---------------------------------------------------------------------------


def pytest_addoption(parser):
    parser.addoption("--pf-email", action="store", default=None, help="PF test user email")
    parser.addoption("--pf-password", action="store", default=None, help="PF test user password")
    parser.addoption("--pf-url", action="store", default=None, help="PF API base URL")
    parser.addoption("--pf-identifier", action="store", default=None, help="PF login identifier")
    parser.addoption("--pf-token", action="store", default=None, help="PF Bearer token (skip login)")
    parser.addoption("--pf-client-id", action="store", default=None, help="PF client ID (with --pf-token)")
    parser.addoption("--pf-customer-id", action="store", default=None, help="PF customer ID (with --pf-token)")
    parser.addoption("--pf-user-id", action="store", default=None, help="PF user ID (with --pf-token)")


# ---------------------------------------------------------------------------
#  Credential resolution
# ---------------------------------------------------------------------------

_DEFAULTS = {
    "email": "ai@mailinator.com",
    "password": "U2FsdGVkX187+r2d+eIWXxAQ2mjzAzhalXdwStSDJP0=",
    "url": "https://api-cx-portal.apps.projectsforce.com",
    "identifier": "projectsforce",
}


def _resolve_credentials(config) -> dict:
    """Resolve credentials from CLI > env > config file > interactive prompt."""
    creds = {}

    # 1. CLI options
    creds["email"] = config.getoption("--pf-email") or ""
    creds["password"] = config.getoption("--pf-password") or ""
    creds["url"] = config.getoption("--pf-url") or ""
    creds["identifier"] = config.getoption("--pf-identifier") or ""

    # 2. Environment variables (fill gaps)
    if not creds["email"]:
        creds["email"] = os.environ.get("PF_TEST_EMAIL", "")
    if not creds["password"]:
        creds["password"] = os.environ.get("PF_TEST_PASSWORD", "")
    if not creds["url"]:
        creds["url"] = os.environ.get("PF_INTEGRATION_URL", "")
    if not creds["identifier"]:
        creds["identifier"] = os.environ.get("PF_TEST_IDENTIFIER", "")

    # 3. Config file
    if os.path.exists(_CREDS_FILE):
        try:
            with open(_CREDS_FILE) as f:
                file_creds = json.load(f)
            for key in ("email", "password", "url", "identifier"):
                if not creds[key] and file_creds.get(key):
                    creds[key] = file_creds[key]
        except (json.JSONDecodeError, OSError):
            pass

    # 4. Interactive prompt (only if stdin is a terminal)
    if sys.stdin.isatty() and (not creds["email"] or not creds["password"]):
        print("\n--- PF Integration Test Credentials ---")
        if not creds["email"]:
            creds["email"] = input(f"  Email [{_DEFAULTS['email']}]: ").strip() or _DEFAULTS["email"]
        if not creds["password"]:
            creds["password"] = input(f"  Password [{_DEFAULTS['password'][:20]}...]: ").strip() or _DEFAULTS["password"]
        if not creds["url"]:
            creds["url"] = input(f"  API URL [{_DEFAULTS['url']}]: ").strip() or _DEFAULTS["url"]
        if not creds["identifier"]:
            creds["identifier"] = input(f"  Identifier [{_DEFAULTS['identifier']}]: ").strip() or _DEFAULTS["identifier"]
        print("---\n")

    # 5. Defaults for anything still empty
    for key, default in _DEFAULTS.items():
        if not creds[key]:
            creds[key] = default

    return creds


# ---------------------------------------------------------------------------
#  HTTP Request/Response Capture
# ---------------------------------------------------------------------------


class HttpCapture:
    """Captures all httpx request/response pairs per test."""

    def __init__(self):
        self.entries: list[dict] = []  # All captured entries
        self._current_test: str = ""
        self._current_question: str = ""

    def set_test(self, test_name: str):
        self._current_test = test_name
        self._current_question = ""

    def set_question(self, question: str):
        self._current_question = question

    def record(self, request: httpx.Request, response: httpx.Response):
        # Parse request body
        req_body = None
        if request.content:
            try:
                req_body = json.loads(request.content.decode("utf-8"))
            except (json.JSONDecodeError, UnicodeDecodeError):
                req_body = request.content.decode("utf-8", errors="replace")

        # Parse response body
        resp_body = None
        try:
            resp_body = response.json()
        except (json.JSONDecodeError, ValueError):
            resp_body = response.text[:5000] if response.text else ""

        # Mask auth token
        auth_header = dict(request.headers).get("authorization", "")
        if auth_header and len(auth_header) > 30:
            auth_header = f"{auth_header[:25]}...{auth_header[-8:]}"

        self.entries.append({
            "test_name": self._current_test,
            "question": self._current_question,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "method": request.method,
            "url": str(request.url),
            "request_headers": {
                k: (auth_header if k.lower() == "authorization" else v)
                for k, v in request.headers.items()
                if k.lower() in ("authorization", "content-type", "accept", "client_id")
            },
            "request_body": req_body,
            "response_status": response.status_code,
            "response_body": resp_body,
        })


# Global capture instance
_capture = HttpCapture()


def _patch_httpx():
    """Monkey-patch httpx.AsyncClient to capture all requests/responses."""
    _original_init = httpx.AsyncClient.__init__

    def _patched_init(self, *args, **kwargs):
        # Inject response event hook
        hooks = kwargs.get("event_hooks", {})
        if "response" not in hooks:
            hooks["response"] = []
        hooks["response"].append(_on_response)
        kwargs["event_hooks"] = hooks
        _original_init(self, *args, **kwargs)

    httpx.AsyncClient.__init__ = _patched_init
    return _original_init


async def _on_response(response: httpx.Response):
    """httpx event hook — called after every response."""
    await response.aread()  # Ensure body is available
    _capture.record(response.request, response)


# ---------------------------------------------------------------------------
#  Excel Report Generation
# ---------------------------------------------------------------------------


def _write_excel_report(entries: list[dict], filepath: str, creds: dict):
    """Write captured HTTP exchanges to an Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["PF Scheduling Bot — Integration Test Report"])
    ws_summary.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["API URL", creds.get("url", "")])
    ws_summary.append(["Test User", creds.get("email", "")])
    ws_summary.append(["Identifier", creds.get("identifier", "")])
    ws_summary.append(["Total API Calls", len(entries)])
    ws_summary.append([])

    # Count by test
    test_counts: dict[str, int] = {}
    for e in entries:
        test_counts[e["test_name"]] = test_counts.get(e["test_name"], 0) + 1
    ws_summary.append(["Test Case", "API Calls"])
    for test, count in test_counts.items():
        ws_summary.append([test, count])

    ws_summary.column_dimensions["A"].width = 55
    ws_summary.column_dimensions["B"].width = 50
    for row in ws_summary.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, size=14)

    # --- Detail sheet ---
    ws = wb.create_sheet("API Calls")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    headers = [
        "#", "Test Case", "Question", "Method", "URL",
        "Request Headers", "Request Body",
        "Response Status", "Response Body",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for idx, entry in enumerate(entries, 1):
        req_headers_str = json.dumps(entry["request_headers"], indent=2) if entry["request_headers"] else ""
        req_body_str = json.dumps(entry["request_body"], indent=2, default=str) if entry["request_body"] else ""
        resp_body_str = (
            json.dumps(entry["response_body"], indent=2, default=str)
            if isinstance(entry["response_body"], (dict, list))
            else str(entry["response_body"] or "")
        )
        # Truncate very long responses for readability
        if len(resp_body_str) > 10000:
            resp_body_str = resp_body_str[:10000] + "\n... (truncated)"

        row = [
            idx,
            entry["test_name"],
            entry.get("question", ""),
            entry["method"],
            entry["url"],
            req_headers_str,
            req_body_str,
            entry["response_status"],
            resp_body_str,
        ]
        ws.append(row)
        for cell in ws[idx + 1]:
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # Column widths
    ws.column_dimensions["A"].width = 5    # #
    ws.column_dimensions["B"].width = 40   # Test Case
    ws.column_dimensions["C"].width = 55   # Question
    ws.column_dimensions["D"].width = 8    # Method
    ws.column_dimensions["E"].width = 70   # URL
    ws.column_dimensions["F"].width = 40   # Request Headers
    ws.column_dimensions["G"].width = 50   # Request Body
    ws.column_dimensions["H"].width = 12   # Response Status
    ws.column_dimensions["I"].width = 80   # Response Body

    # Freeze top row
    ws.freeze_panes = "A2"

    # --- Per-test sheets (grouped) ---
    tests_seen = []
    for e in entries:
        if e["test_name"] not in tests_seen:
            tests_seen.append(e["test_name"])

    for test_name in tests_seen:
        test_entries = [e for e in entries if e["test_name"] == test_name]
        # Sheet name max 31 chars, strip invalid chars for Excel
        raw_name = test_name.split("::")[-1] if "::" in test_name else test_name
        sheet_name = raw_name.replace("[", "_").replace("]", "")[:31]
        ws_test = wb.create_sheet(sheet_name)

        ws_test.append(["Test", test_name])
        ws_test.append(["API Calls", len(test_entries)])
        ws_test.append([])

        for i, entry in enumerate(test_entries, 1):
            ws_test.append([f"--- API Call #{i} ---"])
            ws_test.merge_cells(start_row=ws_test.max_row, start_column=1, end_row=ws_test.max_row, end_column=2)
            ws_test[ws_test.max_row][0].font = Font(bold=True, size=11)

            if entry.get("question"):
                ws_test.append(["Question", entry["question"]])
            ws_test.append(["Method", entry["method"]])
            ws_test.append(["URL", entry["url"]])
            ws_test.append(["Request Headers", json.dumps(entry["request_headers"], indent=2)])
            req_body = json.dumps(entry["request_body"], indent=2, default=str) if entry["request_body"] else "(none)"
            ws_test.append(["Request Body", req_body])
            ws_test.append(["Response Status", entry["response_status"]])
            resp_body = (
                json.dumps(entry["response_body"], indent=2, default=str)
                if isinstance(entry["response_body"], (dict, list))
                else str(entry["response_body"] or "")
            )
            if len(resp_body) > 10000:
                resp_body = resp_body[:10000] + "\n... (truncated)"
            ws_test.append(["Response Body", resp_body])
            ws_test.append([])

        ws_test.column_dimensions["A"].width = 20
        ws_test.column_dimensions["B"].width = 100
        for row in ws_test.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment

    wb.save(filepath)


# ---------------------------------------------------------------------------
#  Session-scoped fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def pf_creds(request):
    """Resolve credentials (CLI > env > file > interactive)."""
    return _resolve_credentials(request.config)


@pytest.fixture(scope="session")
def pf_credentials(pf_creds, request):
    """Authenticate with PF API. Returns credentials dict.

    Supports two modes:
      1. Bearer token (--pf-token / PF_TEST_TOKEN) — skips login, uses token directly
      2. Email/password login (default) — authenticates via PF API
    """
    # Check for direct token (CLI > env > creds file)
    token = request.config.getoption("--pf-token") or os.environ.get("PF_TEST_TOKEN", "")
    if not token and os.path.exists(_CREDS_FILE):
        try:
            with open(_CREDS_FILE) as f:
                file_creds = json.load(f)
            token = file_creds.get("token", "")
        except (json.JSONDecodeError, OSError):
            pass

    if token:
        client_id = (
            request.config.getoption("--pf-client-id")
            or os.environ.get("PF_TEST_CLIENT_ID", "")
            or pf_creds.get("client_id", "16PF11PF")
        )
        customer_id = (
            request.config.getoption("--pf-customer-id")
            or os.environ.get("PF_TEST_CUSTOMER_ID", "")
            or pf_creds.get("customer_id", "90000033")
        )
        user_id = (
            request.config.getoption("--pf-user-id")
            or os.environ.get("PF_TEST_USER_ID", "")
            or pf_creds.get("user_id", customer_id)
        )
        creds = {
            "access_token": token,
            "client_id": client_id,
            "customer_id": str(customer_id),
            "user_id": str(user_id),
            "user_name": "Token User",
            "email": pf_creds.get("email", "token-auth"),
        }
        logger.info(
            "Using direct token auth (client=%s, customer=%s)",
            creds["client_id"], creds["customer_id"],
        )
        return creds

    # Fall back to email/password login
    url = f"{pf_creds['url']}/authentication/login?identifier={pf_creds['identifier']}"
    try:
        resp = httpx.post(
            url,
            json={
                "email": pf_creds["email"],
                "password": pf_creds["password"],
                "device_type": 1,
            },
            timeout=15.0,
        )
    except httpx.ConnectError:
        pytest.skip(f"Cannot reach PF API at {pf_creds['url']}")

    if resp.status_code != 200:
        body = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {"message": resp.text}
        pytest.skip(f"PF login failed ({resp.status_code}): {body.get('message', resp.text[:200])}")

    data = resp.json()
    token = data.get("accesstoken")
    if not token:
        pytest.skip("PF login returned no token")

    user = data.get("user", {})
    creds = {
        "access_token": token,
        "client_id": user.get("client_id", ""),
        "customer_id": str(user.get("customer_id", "")),
        "user_id": str(user.get("customer_id", "")),
        "user_name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip(),
        "email": pf_creds["email"],
    }
    logger.info(
        "Authenticated as %s (client=%s, customer=%s)",
        creds["user_name"], creds["client_id"], creds["customer_id"],
    )
    return creds


@pytest.fixture(scope="session", autouse=True)
def _setup_http_capture():
    """Patch httpx to capture all requests/responses for the report."""
    original_init = _patch_httpx()
    yield
    # Restore
    httpx.AsyncClient.__init__ = original_init


@pytest.fixture(autouse=True)
def _set_auth_context(pf_credentials, pf_creds):
    """Set AuthContext + env vars for every integration test."""
    os.environ["PF_API_BASE_URL"] = pf_creds["url"]
    os.environ["ENVIRONMENT"] = "dev"
    os.environ["USE_DYNAMODB_STORAGE"] = "false"

    from config import get_settings

    get_settings.cache_clear()

    from auth.context import AuthContext

    AuthContext.set(
        auth_token=pf_credentials["access_token"],
        client_id=pf_credentials["client_id"],
        customer_id=pf_credentials["customer_id"],
        user_id=pf_credentials["user_id"],
        user_name=pf_credentials["user_name"],
    )
    yield
    AuthContext.clear()
    get_settings.cache_clear()


@pytest.fixture(autouse=True)
def _clear_caches():
    """Clear tool caches between tests."""
    from tools.scheduling import _projects_cache

    _projects_cache.clear()
    yield
    _projects_cache.clear()


@pytest.fixture(autouse=True)
def _track_test_name(request):
    """Tell the HTTP capture which test is running."""
    _capture.set_test(request.node.nodeid)
    yield


@pytest.fixture()
def http_capture():
    """Access the global HTTP capture to set questions per test."""
    return _capture


# ---------------------------------------------------------------------------
#  Report generation at session end
# ---------------------------------------------------------------------------


def pytest_sessionfinish(session, exitstatus):
    """Write Excel report after all tests complete."""
    if not _capture.entries:
        return

    creds_fixture = session.config._store.get(pytest.StashKey[dict](), {})

    # Build creds summary from env/defaults for the report header
    creds_summary = {
        "url": os.environ.get("PF_API_BASE_URL", _DEFAULTS["url"]),
        "email": os.environ.get("PF_TEST_EMAIL", _DEFAULTS["email"]),
        "identifier": os.environ.get("PF_TEST_IDENTIFIER", _DEFAULTS["identifier"]),
    }

    try:
        _write_excel_report(_capture.entries, _REPORT_FILE, creds_summary)
        print(f"\n{'='*60}")
        print(f"  Report: {_REPORT_FILE}")
        print(f"  API calls captured: {len(_capture.entries)}")
        print(f"{'='*60}")
    except Exception as exc:
        print(f"\nWarning: Failed to write Excel report: {exc}")
