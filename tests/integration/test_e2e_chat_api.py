"""E2E tests via the live /chat API endpoint.

Sends natural-language questions to the deployed chat API and captures the full
HTTP request/response pairs for the Excel report.

Run with:
    AWS_PROFILE=pf-aws uv run pytest tests/integration/test_e2e_chat_api.py -v -s --tb=short

Target endpoint (override with SCHEDULING_BOT_URL env var):
    https://schedulingagent.dev.projectsforce.com/chat
"""

import json
import logging
import os
import uuid
from datetime import datetime, timezone

import httpx
import pytest

logger = logging.getLogger(__name__)

BOT_URL = os.environ.get(
    "SCHEDULING_BOT_URL",
    "https://schedulingagent.dev.projectsforce.com",
)
CHAT_ENDPOINT = f"{BOT_URL}/chat"
_REPORT_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "E2E_FAQ_Test_Report.xlsx")


# ---------------------------------------------------------------------------
#  Report writer — self-contained, writes its own Excel with full req/resp
# ---------------------------------------------------------------------------

_chat_entries: list[dict] = []


def _mask_token(token: str) -> str:
    """Mask auth token for report: show first 10 + last 6 chars."""
    if not token or len(token) < 20:
        return token
    return f"{token[:10]}...{token[-6:]}"


def _write_chat_report(entries: list[dict], filepath: str):
    """Write captured /chat request/response exchanges to an Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    thin = Border(*(Side(style="thin") for _ in range(4)))
    wrap = Alignment(wrap_text=True, vertical="top")
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws.append(["PF Scheduling Bot — E2E Chat API Test Report"])
    ws[1][0].font = Font(bold=True, size=14)
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Endpoint", CHAT_ENDPOINT])
    ws.append(["Total Interactions", len(entries)])
    ws.append([])

    # Group by test class
    groups: dict[str, list[dict]] = {}
    for e in entries:
        groups.setdefault(e["group"], []).append(e)

    ws.append(["Test Group", "Steps", "Result"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for group, items in groups.items():
        all_ok = all(i["status"] == 200 for i in items)
        ws.append([group, len(items), "PASS" if all_ok else "FAIL"])
        ws[ws.max_row][2].fill = pass_fill if all_ok else fail_fill

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12

    # ── Detail sheet ──
    ws2 = wb.create_sheet("All Interactions")
    headers = [
        "#", "Group", "Step", "Question",
        "Request Body", "Response Status",
        "Agent", "Intent", "Bot Response",
    ]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = thin

    for idx, e in enumerate(entries, 1):
        # Mask token in request body for display
        req_display = dict(e["request_body"])
        if "auth_token" in req_display:
            req_display["auth_token"] = _mask_token(req_display["auth_token"])

        resp = e.get("response_body", {}) or {}
        bot_response = resp.get("response", "")
        # Truncate very long responses
        if len(bot_response) > 3000:
            bot_response = bot_response[:3000] + "\n... (truncated)"

        row = [
            idx,
            e["group"],
            e["step"],
            e["question"],
            json.dumps(req_display, indent=2),
            e["status"],
            resp.get("agent_name", ""),
            resp.get("intent", ""),
            bot_response,
        ]
        ws2.append(row)
        for cell in ws2[idx + 1]:
            cell.border = thin
            cell.alignment = wrap

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 25
    ws2.column_dimensions["D"].width = 60
    ws2.column_dimensions["E"].width = 50
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 20
    ws2.column_dimensions["H"].width = 15
    ws2.column_dimensions["I"].width = 100
    ws2.freeze_panes = "A2"

    # ── Per-group sheets with full request + response JSON ──
    for group, items in groups.items():
        sheet_name = group[:31]
        ws_g = wb.create_sheet(sheet_name)
        ws_g.append([f"Test: {group}"])
        ws_g[1][0].font = Font(bold=True, size=12)
        ws_g.append([f"Steps: {len(items)}"])
        ws_g.append([])

        for i, e in enumerate(items, 1):
            ws_g.append([f"--- Step {i}: {e['step']} ---"])
            ws_g[ws_g.max_row][0].font = Font(bold=True, size=11)

            ws_g.append(["Question", e["question"]])

            req_display = dict(e["request_body"])
            if "auth_token" in req_display:
                req_display["auth_token"] = _mask_token(req_display["auth_token"])
            ws_g.append(["Request Body", json.dumps(req_display, indent=2)])

            ws_g.append(["Response Status", e["status"]])

            resp = e.get("response_body", {}) or {}
            resp_display = dict(resp)
            # Truncate response text for readability
            if "response" in resp_display and len(str(resp_display["response"])) > 5000:
                resp_display["response"] = resp_display["response"][:5000] + "... (truncated)"
            ws_g.append(["Response Body", json.dumps(resp_display, indent=2, default=str)])

            ws_g.append(["Agent", resp.get("agent_name", "")])
            ws_g.append(["Intent", resp.get("intent", "")])
            ws_g.append([])

        ws_g.column_dimensions["A"].width = 20
        ws_g.column_dimensions["B"].width = 120
        for row in ws_g.iter_rows():
            for cell in row:
                cell.alignment = wrap

    wb.save(filepath)


# ---------------------------------------------------------------------------
#  Helper: send a chat message, capture, return parsed response
# ---------------------------------------------------------------------------


async def _chat(
    question: str,
    session_id: str,
    creds: dict,
    http_capture,
    *,
    group: str,
    step: str,
) -> dict:
    """Send a message to /chat, print + capture full req/resp, return response."""
    print(f"\n  [{step}] Q: {question}")
    http_capture.set_question(question)

    request_body = {
        "message": question,
        "session_id": session_id,
        "user_id": creds["user_id"],
        "auth_token": creds["access_token"],
        "client_id": creds["client_id"],
        "customer_id": creds["customer_id"],
        "user_name": creds.get("user_name", "Test User"),
    }

    async with httpx.AsyncClient(timeout=90.0) as client:
        resp = await client.post(CHAT_ENDPOINT, json=request_body)

    assert resp.status_code == 200, f"Chat API returned {resp.status_code}: {resp.text[:500]}"
    data = resp.json()

    response_text = data.get("response", "")
    agent = data.get("agent_name", "?")
    intent = data.get("intent", "?")
    preview = response_text[:400].replace("\n", " ")
    print(f"  [{step}] Agent: {agent} | Intent: {intent}")
    print(f"  [{step}] A: {preview}{'...' if len(response_text) > 400 else ''}")

    # Capture for our own report
    _chat_entries.append({
        "group": group,
        "step": step,
        "question": question,
        "request_body": request_body,
        "status": resp.status_code,
        "response_body": data,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })

    return data


# ---------------------------------------------------------------------------
#  Conversation scenarios — ordered, shared session for continuity
# ---------------------------------------------------------------------------

SCENARIOS = [
    {
        "id": "welcome",
        "question": "__WELCOME__",
        "description": "Welcome greeting with project summary",
    },
    {
        "id": "list_all_projects",
        "question": "Show me all my projects",
        "description": "List all customer projects",
    },
    {
        "id": "filter_schedulable",
        "question": "Which of my projects are ready to schedule?",
        "description": "Filter projects by schedulable status",
    },
    {
        "id": "filter_by_category",
        "question": "Show me only my Windows projects",
        "description": "Filter projects by category",
    },
    {
        "id": "filter_scheduled",
        "question": "Which projects are already scheduled?",
        "description": "Filter projects by scheduled status",
    },
    {
        "id": "filter_nonexistent",
        "question": "Do I have any Solar Panel projects?",
        "description": "Filter by non-existent category",
    },
    {
        "id": "project_details",
        "question": "Tell me about my first project — what's the status and address?",
        "description": "Get details for first project",
    },
    {
        "id": "nonexistent_project",
        "question": "What's the status of project 99999999?",
        "description": "Lookup non-existent project ID",
    },
    {
        "id": "business_hours",
        "question": "What are your business hours?",
        "description": "Get business hours",
    },
    {
        "id": "project_address",
        "question": "What is the installation address for my projects?",
        "description": "Get project installation addresses",
    },
    {
        "id": "project_installer",
        "question": "Who is the technician assigned to my projects?",
        "description": "Get installer/technician info",
    },
    {
        "id": "store_info",
        "question": "Which store is handling my projects?",
        "description": "Get store info for projects",
    },
    {
        "id": "weather",
        "question": "What's the weather like at my project's location?",
        "description": "Get weather for project address",
    },
    {
        "id": "add_note",
        "question": "Add a note to my first project: Customer requested morning appointment",
        "description": "Add a note to a project",
    },
    {
        "id": "list_notes",
        "question": "Show me the notes for that project",
        "description": "List notes for a project",
    },
    {
        "id": "all_info_incorrect",
        "question": "All my project information is showing incorrectly, can you check?",
        "description": "Customer complaint — verify project data",
    },
    {
        "id": "greeting",
        "question": "Hello! How are you?",
        "description": "Casual greeting (chitchat routing)",
    },
    {
        "id": "off_topic",
        "question": "What's the capital of France?",
        "description": "Off-topic question (guardrail test)",
    },
]


# ---------------------------------------------------------------------------
#  1. General scenario tests (single-turn)
# ---------------------------------------------------------------------------


class TestChatAPI:
    """Send each scenario question to the live /chat API."""

    _session_id: str = ""

    @pytest.fixture(autouse=True, scope="class")
    def _init_session(self):
        TestChatAPI._session_id = str(uuid.uuid4())
        yield

    @pytest.mark.parametrize("scenario", SCENARIOS, ids=[s["id"] for s in SCENARIOS])
    async def test_chat_scenario(self, scenario, pf_credentials, http_capture):
        data = await _chat(
            scenario["question"],
            self._session_id,
            pf_credentials,
            http_capture,
            group="General Scenarios",
            step=scenario["id"],
        )
        assert "response" in data
        assert "session_id" in data
        assert len(data.get("response", "")) > 5


# ---------------------------------------------------------------------------
#  2. Schedule Balcony Grill (Order #74356) — full multi-turn flow
# ---------------------------------------------------------------------------


class TestScheduleFenceInstallation:
    """Full scheduling flow for Fence Installation project (Order #74356_1).

    Flow: list schedulable -> pick Fence Installation #74356_1 -> get dates ->
          pick date -> get time slots -> pick slot -> confirm -> cancel (cleanup)
    """

    async def test_schedule_flow(self, pf_credentials, http_capture):
        session_id = str(uuid.uuid4())
        creds = pf_credentials
        G = "Schedule Fence Installation #74356_1"

        # Step 1: List schedulable projects
        r = await _chat(
            "Show me my projects that are ready to schedule",
            session_id, creds, http_capture,
            group=G, step="1. List schedulable",
        )
        assert len(r["response"]) > 20

        # Step 2: Pick Fence Installation by order number and ask for dates
        r = await _chat(
            "I want to schedule my Fence Installation project, order number 74356_1. What dates are available?",
            session_id, creds, http_capture,
            group=G, step="2. Get available dates",
        )
        assert len(r["response"]) > 20

        # Step 3: Ask for time slots on first date
        r = await _chat(
            "Show me the time slots for the first available date",
            session_id, creds, http_capture,
            group=G, step="3. Get time slots",
        )
        assert len(r["response"]) > 10

        # Step 4: Pick the first time slot
        r = await _chat(
            "I'll take the first available time slot please",
            session_id, creds, http_capture,
            group=G, step="4. Pick time slot",
        )
        assert len(r["response"]) > 10

        # Step 5: Confirm the appointment
        r = await _chat(
            "Yes, go ahead and schedule it",
            session_id, creds, http_capture,
            group=G, step="5. Confirm appointment",
        )
        assert len(r["response"]) > 10

        # Step 6: Cancel the appointment (cleanup so test is idempotent)
        r = await _chat(
            "Actually, please cancel that appointment",
            session_id, creds, http_capture,
            group=G, step="6. Cancel (cleanup)",
        )
        assert len(r["response"]) > 10


# ---------------------------------------------------------------------------
#  3. Reschedule Windows (Order #6789) — full multi-turn flow
# ---------------------------------------------------------------------------


class TestRescheduleWindows:
    """Full reschedule flow for Windows project (Order #6789).

    Flow: ask to reschedule order 6789 -> pick new date -> get slots ->
          pick slot -> confirm new appointment
    """

    async def test_reschedule_flow(self, pf_credentials, http_capture):
        session_id = str(uuid.uuid4())
        creds = pf_credentials
        G = "Reschedule Windows #6789"

        # Step 1: Ask to reschedule Windows order 6789
        r = await _chat(
            "I need to reschedule my Windows project, order number 6789",
            session_id, creds, http_capture,
            group=G, step="1. Request reschedule",
        )
        assert len(r["response"]) > 20

        # Step 2: Confirm we want to reschedule
        r = await _chat(
            "Yes, go ahead and reschedule it",
            session_id, creds, http_capture,
            group=G, step="2. Confirm reschedule",
        )
        assert len(r["response"]) > 10

        # Step 3: Pick a new date
        r = await _chat(
            "I'll take the first available date",
            session_id, creds, http_capture,
            group=G, step="3. Pick new date",
        )
        assert len(r["response"]) > 10

        # Step 4: Get time slots
        r = await _chat(
            "What time slots are available?",
            session_id, creds, http_capture,
            group=G, step="4. Get time slots",
        )
        assert len(r["response"]) > 10

        # Step 5: Pick the first slot
        r = await _chat(
            "I'll take the first slot",
            session_id, creds, http_capture,
            group=G, step="5. Pick time slot",
        )
        assert len(r["response"]) > 10

        # Step 6: Confirm the new appointment
        r = await _chat(
            "Yes, confirm the new appointment",
            session_id, creds, http_capture,
            group=G, step="6. Confirm new appointment",
        )
        assert len(r["response"]) > 10


# ---------------------------------------------------------------------------
#  Write our own report at session end (via autouse session fixture)
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session", autouse=True)
def _write_report_on_finish():
    """Write the E2E Chat API report after all tests complete."""
    yield
    if not _chat_entries:
        return
    try:
        _write_chat_report(_chat_entries, _REPORT_FILE)
        print(f"\n{'='*60}")
        print(f"  E2E Chat API Report: {_REPORT_FILE}")
        print(f"  Interactions captured: {len(_chat_entries)}")
        print(f"{'='*60}")
    except Exception as exc:
        print(f"\nWarning: Failed to write report: {exc}")
